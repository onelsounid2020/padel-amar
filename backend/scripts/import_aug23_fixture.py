from __future__ import annotations

import argparse
import json
import os
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from urllib.error import HTTPError
from urllib.request import Request, urlopen

from openpyxl import load_workbook


EVENT_ID = 18
EVENT_DATE = "2026-08-23"
DEFAULT_WORKBOOK = Path("/Users/onel/Downloads/sunday_23-agosto_7canchas_MF.xlsx")
DEFAULT_API_URL = "https://radiant-warmth-production-a8ec.up.railway.app"

SCHEDULES = [
    {
        "sheet": "MALE",
        "category": "Hombres",
        "group": "Grupo A",
        "pair_cells": ["B7", "B8", "B9", "B10", "B11", "B12"],
        "court_row": 6,
        "court_cols": ["H", "J", "L"],
        "time_col": "F",
        "rows": [7, 9, 11, 13, 15],
    },
    {
        "sheet": "female",
        "category": "Mujeres",
        "group": "Grupo A",
        "pair_cells": ["B9", "B10", "B11", "B12"],
        "court_row": 8,
        "court_cols": ["J", "M"],
        "time_col": "H",
        "rows": [9, 11, 13],
    },
    {
        "sheet": "female",
        "category": "Mujeres",
        "group": "Grupo B",
        "pair_cells": ["B16", "B17", "B18", "B19"],
        "court_row": 15,
        "court_cols": ["J", "M"],
        "time_col": "H",
        "rows": [16, 18, 20],
    },
]

NAME_FIXES = {
    "nathy": "Nathy",
    "andrea": "Andrea",
}


def clean(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def clean_name(value: str) -> str:
    value = clean(value)
    return NAME_FIXES.get(value.lower(), value[:1].upper() + value[1:] if value else value)


def split_pair(label: str) -> tuple[str, str]:
    parts = [clean_name(part) for part in clean(label).split("/", 1)]
    if len(parts) != 2 or not all(parts):
        raise ValueError(f"No pude separar la pareja: {label!r}")
    return parts[0], parts[1]


def pair_key(label: str, category: str) -> str:
    one, two = split_pair(label)
    return f"{category.lower()}::{one.lower()}::{two.lower()}"


def normalize_court(value: object) -> str:
    court = clean(value)
    match = re.search(r"(\d+)", court)
    if not match:
        raise ValueError(f"Cancha inválida: {value!r}")
    return match.group(1)


def normalize_slot(value: object) -> str:
    slot = clean(value).replace(".", ":").replace(" - ", "-")
    slot = re.sub(r"\s+", "", slot)
    if not re.fullmatch(r"\d{1,2}:\d{2}-\d{1,2}:\d{2}", slot):
        raise ValueError(f"Horario inválido: {value!r}")
    return slot


def parse_workbook(path: Path) -> tuple[list[dict], list[dict]]:
    workbook = load_workbook(path, data_only=True)
    if workbook.sheetnames != ["MALE", "female"]:
        raise ValueError(f"Se esperaban exactamente las hojas MALE y female; recibí {workbook.sheetnames}")

    pairs: list[dict] = []
    matches: list[dict] = []
    for schedule in SCHEDULES:
        sheet = workbook[schedule["sheet"]]
        for cell in schedule["pair_cells"]:
            label = clean(sheet[cell].value)
            one, two = split_pair(label)
            pairs.append({
                "key": pair_key(label, schedule["category"]),
                "one": one,
                "two": two,
                "category": schedule["category"],
                "group": schedule["group"],
            })
        for round_number, row in enumerate(schedule["rows"], start=1):
            slot = normalize_slot(sheet[f"{schedule['time_col']}{row}"].value)
            for court_col in schedule["court_cols"]:
                one_label = clean(sheet[f"{court_col}{row}"].value)
                two_label = clean(sheet[f"{court_col}{row + 1}"].value)
                matches.append({
                    "one_key": pair_key(one_label, schedule["category"]),
                    "two_key": pair_key(two_label, schedule["category"]),
                    "category": schedule["category"],
                    "group": schedule["group"],
                    "round_name": f"{schedule['category']} - {schedule['group']} - Ronda {round_number} - {slot}",
                    "court": normalize_court(sheet[f"{court_col}{schedule['court_row']}"].value),
                    "slot": slot,
                })

    pair_keys = {pair["key"] for pair in pairs}
    if len(pairs) != 14 or len(pair_keys) != 14:
        raise ValueError(f"Se esperaban 14 parejas únicas; recibí {len(pairs)} filas y {len(pair_keys)} únicas")
    if len(matches) != 27:
        raise ValueError(f"Se esperaban 27 partidos; recibí {len(matches)}")
    if any(match["one_key"] not in pair_keys or match["two_key"] not in pair_keys for match in matches):
        raise ValueError("Hay partidos con parejas fuera de las listas de las dos hojas")

    schedule_keys = [(match["slot"], match["court"]) for match in matches]
    if len(schedule_keys) != len(set(schedule_keys)):
        duplicates = [key for key, count in Counter(schedule_keys).items() if count > 1]
        raise ValueError(f"Hay canchas duplicadas en un mismo horario: {duplicates}")

    pair_courts: dict[str, set[str]] = defaultdict(set)
    pair_games: Counter[str] = Counter()
    for match in matches:
        for key in (match["one_key"], match["two_key"]):
            pair_courts[key].add(match["court"])
            pair_games[key] += 1

    fixed = {
        "hombres::onel::arturo": ("2", 5),
        "mujeres::nathy::judith": ("1", 3),
    }
    for key, (court, games) in fixed.items():
        if pair_courts[key] != {court} or pair_games[key] != games:
            raise ValueError(f"La pareja prioritaria {key} no quedó fija: canchas={pair_courts[key]}, partidos={pair_games[key]}")

    expected_games = {pair["key"]: (5 if pair["category"] == "Hombres" else 3) for pair in pairs}
    if dict(pair_games) != expected_games:
        raise ValueError("La cantidad de partidos por pareja no coincide con el formato de las hojas")
    return pairs, matches


class Api:
    def __init__(self, base_url: str, token: str | None = None):
        self.base_url = base_url.rstrip("/")
        self.token = token

    def request(self, path: str, method: str = "GET", payload: dict | None = None):
        headers = {"Content-Type": "application/json"}
        if self.token:
            headers["Authorization"] = f"Bearer {self.token}"
        request = Request(
            f"{self.base_url}{path}",
            data=json.dumps(payload).encode() if payload is not None else None,
            headers=headers,
            method=method,
        )
        try:
            with urlopen(request, timeout=45) as response:
                body = response.read()
                return json.loads(body) if body else None
        except HTTPError as error:
            raise RuntimeError(f"{method} {path}: HTTP {error.code} {error.read().decode()}") from error


def snapshot(api: Api, backup_dir: Path) -> Path:
    payload = {
        "created_at": datetime.now().isoformat(),
        "event": api.request(f"/events/{EVENT_ID}"),
        "pairs": api.request(f"/events/{EVENT_ID}/pairs"),
        "matches": api.request(f"/events/{EVENT_ID}/matches"),
        "standings": api.request(f"/events/{EVENT_ID}/standings"),
    }
    backup_dir.mkdir(parents=True, exist_ok=True)
    path = backup_dir / f"event-{EVENT_ID}-backup-before-aug23-fixture-{datetime.now():%Y%m%d-%H%M%S}.json"
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--api-url", default=DEFAULT_API_URL)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--backup-dir", type=Path, default=Path(__file__).resolve().parents[2] / "tmp")
    args = parser.parse_args()

    pairs, matches = parse_workbook(args.workbook)
    print("Hojas revisadas: MALE y female")
    print(f"Parejas: {len(pairs)} ({Counter(pair['category'] for pair in pairs)})")
    print(f"Partidos: {len(matches)} ({Counter(match['category'] for match in matches)})")
    print("Onel/Arturo: 5 partidos en cancha 2")
    print("Nathy/Judith: 3 partidos en cancha 1")
    if not args.apply:
        print("Dry run correcto. Usa --apply para cargar el evento.")
        return

    email = os.environ.get("AMAR_ADMIN_EMAIL")
    password = os.environ.get("AMAR_ADMIN_PASSWORD")
    if not email or not password:
        raise ValueError("Faltan AMAR_ADMIN_EMAIL y AMAR_ADMIN_PASSWORD")

    api = Api(args.api_url)
    auth = api.request("/auth/login", "POST", {"email": email, "password": password})
    api.token = auth["access_token"]
    event = api.request(f"/events/{EVENT_ID}")
    if event["date"] != EVENT_DATE:
        raise ValueError(f"El evento {EVENT_ID} no corresponde al {EVENT_DATE}")
    existing_pairs = api.request(f"/events/{EVENT_ID}/pairs")
    existing_matches = api.request(f"/events/{EVENT_ID}/matches")
    if existing_pairs or existing_matches:
        raise ValueError(f"El evento no está vacío: {len(existing_pairs)} parejas y {len(existing_matches)} partidos")
    backup = snapshot(api, args.backup_dir)

    fixture_config = {
        **(event.get("fixture_config") or {}),
        "mode": "groups",
        "group_size": 4,
        "guaranteed_matches": 3,
        "court_count": 7,
        "courts": "1, 2, 3, 4, 5, 6, 7",
        "start_time": "15:40",
        "set_minutes": 25,
        "category_playoff_modes": {"Mujeres": "crossed_semifinals"},
    }
    category_configs = [
        {"category": "Hombres", "modality": "groups", "group_size": 6, "guaranteed_matches": 5, "qualifiers_per_group": 0, "notes": "Grupo A; Onel/Arturo fijos en cancha 2."},
        {"category": "Mujeres", "modality": "groups_crossed_semifinals", "group_size": 4, "guaranteed_matches": 3, "qualifiers_per_group": 2, "notes": "Grupos A y B; clasifican dos por grupo a semifinales cruzadas; Nathy/Judith fijas en cancha 1."},
    ]
    api.request(f"/events/{EVENT_ID}", "PATCH", {
        "capacity": 14,
        "categories": "Hombres / Mujeres",
        "schedule": "15:40 - 17:45",
        "category_configs": category_configs,
        "fixture_config": fixture_config,
        "fixture_visible": True,
        "status": "registration_closed",
    })

    players = api.request("/players")
    players_by_key = {(clean(player["name"]).lower(), player["category"]): player for player in players}

    def get_or_create_player(name: str, category: str) -> dict:
        key = (name.lower(), category)
        if key not in players_by_key:
            players_by_key[key] = api.request("/players", "POST", {
                "name": name,
                "email": None,
                "phone": None,
                "category": category,
                "preferred_side": "indiferente",
            })
        return players_by_key[key]

    pairs_by_key: dict[str, dict] = {}
    for seed, item in enumerate(pairs, start=1):
        one = get_or_create_player(item["one"], item["category"])
        two = get_or_create_player(item["two"], item["category"])
        pairs_by_key[item["key"]] = api.request(f"/events/{EVENT_ID}/pairs", "POST", {
            "player_one_id": one["id"],
            "player_two_id": two["id"],
            "category": item["category"],
            "skill_level": 5,
            "status": "completa",
            "seed": seed,
        })

    created = api.request(f"/events/{EVENT_ID}/matches/bulk", "POST", {
        "matches": [{
            "pair_one_id": pairs_by_key[match["one_key"]]["id"],
            "pair_two_id": pairs_by_key[match["two_key"]]["id"],
            "round_name": match["round_name"],
            "court": match["court"],
            "played_at": None,
        } for match in matches],
        "replace_unplayed": False,
        "replace_category": None,
    })
    api.request(f"/events/{EVENT_ID}/standings/recalculate", "POST", {})
    print(f"Importación completa: {len(pairs_by_key)} parejas y {len(created)} partidos. Respaldo: {backup}")


if __name__ == "__main__":
    main()
